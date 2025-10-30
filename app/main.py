import pymysql
from openpyxl import Workbook
from datetime import datetime
import os

# 数据库配置
DB_CONFIG = {'host': 'rr-bp102335g97l5k6318o.mysql.rds.aliyuncs.com', 'user': 'jonny', 'password': 'kuang123BYPMS', 'database': 'dh_short_rent'}


def get_db_connection():
    """数据库连接函数"""
    return pymysql.connect(**DB_CONFIG)


def fetch_stats(start_date, end_date):
    """从数据库中获取统计数据"""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            sql = """
            SELECT 
                stats.date,
                stats.hour,
                SUM(stats.order_count) AS order_count,
                SUM(stats.checkin_count) AS checkin_count,
                SUM(stats.checkout_count) AS checkout_count
            FROM (
                -- 订单数据
                SELECT 
                    DATE(create_time) AS date,
                    HOUR(create_time) AS hour,
                    COUNT(*) AS order_count,
                    0 AS checkin_count,
                    0 AS checkout_count
                FROM contract 
                WHERE company_id = %s 
                AND state <> 'D'
                AND DATE(create_time) BETWEEN %s AND %s
                GROUP BY DATE(create_time), HOUR(create_time)

                UNION ALL

                -- 入住数据
                SELECT 
                    DATE(create_time) AS date,
                    HOUR(create_time) AS hour,
                    0 AS order_count,
                    COUNT(*) AS checkin_count,
                    0 AS checkout_count
                FROM empl_action_log
                WHERE company_id = %s
                AND content LIKE %s
                AND DATE(create_time) BETWEEN %s AND %s
                GROUP BY DATE(create_time), HOUR(create_time)

                UNION ALL

                -- 离店数据
                SELECT 
                    DATE(create_time) AS date,
                    HOUR(create_time) AS hour,
                    0 AS order_count,
                    0 AS checkin_count,
                    COUNT(*) AS checkout_count
                FROM empl_action_log
                WHERE company_id = %s 
                AND content LIKE %s
                AND DATE(create_time) BETWEEN %s AND %s
                GROUP BY DATE(create_time), HOUR(create_time)
            ) AS stats
            GROUP BY stats.date, stats.hour
            ORDER BY stats.date, stats.hour;
            """
            company_id = "163924"
            cursor.execute(sql, (company_id, start_date, end_date, company_id, "%已入住房间%", start_date, end_date, company_id, "%已退房%", start_date, end_date))
            return cursor.fetchall()
    finally:
        connection.close()


def generate_excel_report(start_date, end_date, output_file):
    """生成 Excel 报告"""
    # 获取统计数据
    raw_data = fetch_stats(start_date, end_date)

    # 初始化数据结构
    hours = list(range(24))  # 小时范围：0-23
    dates = sorted(set(row[0] for row in raw_data))  # 去重并排序的日期列表
    data = {"order_count": {date: {hour: 0 for hour in hours} for date in dates}, "checkin_count": {date: {hour: 0 for hour in hours} for date in dates},
        "checkout_count": {date: {hour: 0 for hour in hours} for date in dates}}

    # 填充数据
    for row in raw_data:
        date, hour, order_count, checkin_count, checkout_count = row
        data["order_count"][date][hour] = order_count
        data["checkin_count"][date][hour] = checkin_count
        data["checkout_count"][date][hour] = checkout_count

    # 创建 Excel 文件
    wb = Workbook()
    sheets = {"订单统计": wb.create_sheet("订单统计"), "入住统计": wb.create_sheet("入住统计"), "离店统计": wb.create_sheet("离店统计")}

    # 删除默认 Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 中文 Sheet 名称到英文键名的映射
    sheet_to_key = {"订单统计": "order_count", "入住统计": "checkin_count", "离店统计": "checkout_count"}

    # 写入表头和数据
    for sheet_name, sheet in sheets.items():
        # 表头：第一行为日期，第一列为小时
        header = ["小时"] + [d.strftime('%Y-%m-%d') for d in dates]
        sheet.append(header)

        # 写入每小时的数据
        for hour in hours:
            row_data = [str(hour)] + [data[sheet_to_key[sheet_name]][date][hour] for date in dates]
            sheet.append(row_data)

    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 保存文件
    wb.save(output_file)
    print(f"报表已生成到: {output_file}")


if __name__ == '__main__':
    # 指定完整文件路径
    output_file = r"D:\报表输出\酒店每日统计1.xlsx"
    generate_excel_report(start_date="2025-01-27", end_date="2025-03-26", output_file=output_file)